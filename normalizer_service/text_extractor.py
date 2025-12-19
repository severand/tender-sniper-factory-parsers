"""Text extraction from various file formats"""

import os
import tempfile
from typing import Optional
from pathlib import Path

fromshared.logger import logger


class TextExtractor:
    """Extract text from various file formats"""
    
    @staticmethod
    def extract_from_pdf(file_path: str) -> str:
        """Extract text from PDF
        
        Args:
            file_path: Path to PDF file
        
        Returns:
            Extracted text
        """
        try:
            import PyPDF2
            
            text = []
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text.append(page.extract_text())
            
            result = '\n'.join(text)
            logger.info(f"Extracted {len(result)} chars from PDF: {file_path}")
            return result
        except Exception as e:
            logger.error(f"Failed to extract PDF: {str(e)}")
            return ""
    
    @staticmethod
    def extract_from_docx(file_path: str) -> str:
        """Extract text from DOCX
        
        Args:
            file_path: Path to DOCX file
        
        Returns:
            Extracted text
        """
        try:
            from docx import Document
            
            doc = Document(file_path)
            text = [paragraph.text for paragraph in doc.paragraphs]
            
            result = '\n'.join(text)
            logger.info(f"Extracted {len(result)} chars from DOCX: {file_path}")
            return result
        except Exception as e:
            logger.error(f"Failed to extract DOCX: {str(e)}")
            return ""
    
    @staticmethod
    def extract_from_xlsx(file_path: str) -> str:
        """Extract text from XLSX
        
        Args:
            file_path: Path to XLSX file
        
        Returns:
            Extracted text
        """
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(file_path)
            text = []
            
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    text.extend([str(cell) for cell in row if cell])
            
            result = '\n'.join(text)
            logger.info(f"Extracted {len(result)} chars from XLSX: {file_path}")
            return result
        except Exception as e:
            logger.error(f"Failed to extract XLSX: {str(e)}")
            return ""
    
    @staticmethod
    def extract_from_txt(file_path: str) -> str:
        """Extract text from TXT
        
        Args:
            file_path: Path to TXT file
        
        Returns:
            Extracted text
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                text = file.read()
            logger.info(f"Extracted {len(text)} chars from TXT: {file_path}")
            return text
        except Exception as e:
            logger.error(f"Failed to extract TXT: {str(e)}")
            return ""
    
    @staticmethod
    def extract_from_file(file_path: str) -> str:
        """Extract text from file (auto-detect format)
        
        Args:
            file_path: Path to file
        
        Returns:
            Extracted text
        """
        path = Path(file_path)
        suffix = path.suffix.lower()
        
        logger.info(f"Extracting text from {suffix} file: {file_path}")
        
        if suffix == '.pdf':
            return TextExtractor.extract_from_pdf(file_path)
        elif suffix == '.docx':
            return TextExtractor.extract_from_docx(file_path)
        elif suffix == '.xlsx':
            return TextExtractor.extract_from_xlsx(file_path)
        elif suffix == '.txt':
            return TextExtractor.extract_from_txt(file_path)
        else:
            logger.warning(f"Unsupported file format: {suffix}")
            return ""
    
    @staticmethod
    def extract_from_url(url: str) -> str:
        """Extract text from URL
        
        Args:
            url: URL to file
        
        Returns:
            Extracted text
        """
        try:
            import requests
            
            response = requests.get(url, timeout=30)
            response.raise_for_status()
            
            # Save to temp file
            with tempfile.NamedTemporaryFile(delete=False, suffix=Path(url).suffix) as tmp:
                tmp.write(response.content)
                tmp_path = tmp.name
            
            # Extract from temp file
            text = TextExtractor.extract_from_file(tmp_path)
            
            # Cleanup
            os.unlink(tmp_path)
            
            logger.info(f"Extracted text from URL: {url}")
            return text
        except Exception as e:
            logger.error(f"Failed to extract from URL: {str(e)}")
            return ""
